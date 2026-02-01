"""输出生成器

支持生成多种格式的交付物文件。
"""

from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
import json


@dataclass
class OutputConfig:
    """输出配置"""
    title: str = ""
    author: str = "AI Auto"
    template: Optional[str] = None
    styles: dict = field(default_factory=dict)
    metadata: dict = field(default_factory=dict)


class OutputGenerator(ABC):
    """输出生成器基类"""
    
    @property
    @abstractmethod
    def format_name(self) -> str:
        """格式名称"""
        pass
    
    @property
    @abstractmethod
    def file_extension(self) -> str:
        """文件扩展名"""
        pass
    
    @abstractmethod
    async def generate(
        self,
        data: Any,
        output_path: Path,
        config: Optional[OutputConfig] = None,
    ) -> Path:
        """生成输出文件"""
        pass


class ExcelGenerator(OutputGenerator):
    """Excel 生成器"""
    
    @property
    def format_name(self) -> str:
        return "Excel"
    
    @property
    def file_extension(self) -> str:
        return ".xlsx"
    
    async def generate(
        self,
        data: Any,
        output_path: Path,
        config: Optional[OutputConfig] = None,
    ) -> Path:
        """生成 Excel 文件
        
        Args:
            data: 可以是以下格式:
                - list[dict]: 表格数据
                - dict[str, list[dict]]: 多个工作表
                - pandas.DataFrame: 数据框
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError("需要安装: pip install pandas openpyxl")
        
        config = config or OutputConfig()
        output_path = Path(output_path)
        
        if not output_path.suffix:
            output_path = output_path.with_suffix(self.file_extension)
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        
        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        def write_sheet(ws, sheet_data: list[dict], sheet_name: str = None):
            """写入工作表"""
            if sheet_name:
                ws.title = sheet_name
            
            if not sheet_data:
                return
            
            # 写入表头
            headers = list(sheet_data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 写入数据
            for row_idx, row_data in enumerate(sheet_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                    cell.border = thin_border
            
            # 调整列宽
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row in sheet_data:
                    val = str(row.get(header, ""))
                    max_length = max(max_length, len(val))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # 处理不同类型的数据
        if isinstance(data, pd.DataFrame):
            ws = wb.active
            ws.title = config.title or "Sheet1"
            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
        
        elif isinstance(data, dict):
            # 多个工作表
            ws = wb.active
            first = True
            for sheet_name, sheet_data in data.items():
                if first:
                    write_sheet(ws, sheet_data, sheet_name)
                    first = False
                else:
                    new_ws = wb.create_sheet(title=sheet_name)
                    write_sheet(new_ws, sheet_data, sheet_name)
        
        elif isinstance(data, list):
            ws = wb.active
            ws.title = config.title or "Sheet1"
            write_sheet(ws, data)
        
        wb.save(str(output_path))
        return output_path


class MarkdownGenerator(OutputGenerator):
    """Markdown 生成器"""
    
    @property
    def format_name(self) -> str:
        return "Markdown"
    
    @property
    def file_extension(self) -> str:
        return ".md"
    
    async def generate(
        self,
        data: Any,
        output_path: Path,
        config: Optional[OutputConfig] = None,
    ) -> Path:
        """生成 Markdown 文件
        
        Args:
            data: 可以是以下格式:
                - str: 直接文本
                - dict: 包含 title, content, sections 等
                - list[dict]: 表格数据
        """
        config = config or OutputConfig()
        output_path = Path(output_path)
        
        if not output_path.suffix:
            output_path = output_path.with_suffix(self.file_extension)
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        content = ""
        
        if isinstance(data, str):
            content = data
        
        elif isinstance(data, dict):
            # 标题
            if "title" in data or config.title:
                content += f"# {data.get('title', config.title)}\n\n"
            
            # 元数据
            if "metadata" in data:
                for key, value in data["metadata"].items():
                    content += f"**{key}**: {value}  \n"
                content += "\n---\n\n"
            
            # 正文
            if "content" in data:
                content += data["content"] + "\n\n"
            
            # 章节
            if "sections" in data:
                for section in data["sections"]:
                    level = section.get("level", 2)
                    content += f"{'#' * level} {section.get('title', '')}\n\n"
                    content += section.get("content", "") + "\n\n"
        
        elif isinstance(data, list) and data and isinstance(data[0], dict):
            # 表格数据
            if config.title:
                content += f"# {config.title}\n\n"
            
            headers = list(data[0].keys())
            content += "| " + " | ".join(headers) + " |\n"
            content += "| " + " | ".join(["---"] * len(headers)) + " |\n"
            
            for row in data:
                values = [str(row.get(h, "")) for h in headers]
                content += "| " + " | ".join(values) + " |\n"
        
        output_path.write_text(content, encoding="utf-8")
        return output_path


class PDFGenerator(OutputGenerator):
    """PDF 生成器"""
    
    @property
    def format_name(self) -> str:
        return "PDF"
    
    @property
    def file_extension(self) -> str:
        return ".pdf"
    
    async def generate(
        self,
        data: Any,
        output_path: Path,
        config: Optional[OutputConfig] = None,
    ) -> Path:
        """生成 PDF 文件
        
        Args:
            data: 可以是以下格式:
                - str: Markdown 或 HTML 文本
                - dict: 包含 content, styles 等
                - Path: Markdown 文件路径
        """
        config = config or OutputConfig()
        output_path = Path(output_path)
        
        if not output_path.suffix:
            output_path = output_path.with_suffix(self.file_extension)
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 获取内容
        content = ""
        if isinstance(data, str):
            content = data
        elif isinstance(data, dict):
            content = data.get("content", "")
        elif isinstance(data, Path) and data.exists():
            content = data.read_text(encoding="utf-8")
        
        # 转换 Markdown 为 HTML
        try:
            import markdown
            html_content = markdown.markdown(
                content,
                extensions=["tables", "fenced_code", "codehilite"],
            )
        except ImportError:
            html_content = f"<pre>{content}</pre>"
        
        # 添加样式
        styled_html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{config.title or 'Document'}</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            line-height: 1.6;
            max-width: 800px;
            margin: 0 auto;
            padding: 20px;
            color: #333;
        }}
        h1, h2, h3 {{ color: #2563eb; }}
        table {{
            border-collapse: collapse;
            width: 100%;
            margin: 1em 0;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }}
        th {{
            background-color: #4472C4;
            color: white;
        }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        code {{
            background-color: #f4f4f4;
            padding: 2px 6px;
            border-radius: 4px;
        }}
        pre {{
            background-color: #f4f4f4;
            padding: 16px;
            border-radius: 8px;
            overflow-x: auto;
        }}
    </style>
</head>
<body>
{html_content}
</body>
</html>"""
        
        # 尝试使用 weasyprint 生成 PDF
        try:
            from weasyprint import HTML
            HTML(string=styled_html).write_pdf(str(output_path))
        except ImportError:
            # 回退：保存为 HTML
            html_path = output_path.with_suffix(".html")
            html_path.write_text(styled_html, encoding="utf-8")
            return html_path
        
        return output_path


class JSONGenerator(OutputGenerator):
    """JSON 生成器"""
    
    @property
    def format_name(self) -> str:
        return "JSON"
    
    @property
    def file_extension(self) -> str:
        return ".json"
    
    async def generate(
        self,
        data: Any,
        output_path: Path,
        config: Optional[OutputConfig] = None,
    ) -> Path:
        """生成 JSON 文件"""
        config = config or OutputConfig()
        output_path = Path(output_path)
        
        if not output_path.suffix:
            output_path = output_path.with_suffix(self.file_extension)
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 添加元数据
        if config.metadata or config.title:
            wrapped_data = {
                "metadata": {
                    "title": config.title,
                    "author": config.author,
                    "generated_at": datetime.now().isoformat(),
                    **config.metadata,
                },
                "data": data,
            }
        else:
            wrapped_data = data
        
        output_path.write_text(
            json.dumps(wrapped_data, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        return output_path


# 生成器注册表
_generators: dict[str, OutputGenerator] = {}


def register_generator(generator: OutputGenerator) -> None:
    """注册生成器"""
    _generators[generator.format_name.lower()] = generator
    _generators[generator.file_extension.lower()] = generator


def get_output_generator(format_or_ext: str) -> Optional[OutputGenerator]:
    """获取生成器"""
    key = format_or_ext.lower()
    if not key.startswith("."):
        key_with_dot = f".{key}"
    else:
        key_with_dot = key
        key = key[1:]
    
    return _generators.get(key) or _generators.get(key_with_dot)


# 注册内置生成器
register_generator(ExcelGenerator())
register_generator(MarkdownGenerator())
register_generator(PDFGenerator())
register_generator(JSONGenerator())
